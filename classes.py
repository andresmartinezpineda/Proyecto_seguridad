# Importar módulos estándar y clases auxiliares
from openpyxl import load_workbook # manipulación de archivos Excel (.xlsx)
import sys                      # redirigir stdout/stderr cuando sea necesario
import os                       # operaciones de sistema de archivos (path, exists, listdir, makedirs, etc.)
import shutil                   # funciones para copiar archivos preservando metadatos (copy2)
import calendar                 # utilidades para nombres de meses y formatos relacionados
from slack_bot import NOTIFIER  # notificador de Slack preconfigurado (puede ser None)
from datetime import datetime  # obtener fecha y hora actuales


BASE_DIR = os.path.dirname(os.path.abspath(__file__)) # Directorio base del proyecto
CONFIG_PATH = os.path.join(BASE_DIR, "config", "settings.xlsx") # Ruta al archivo de configuración Excel

def _load_config_sheet(): 
    wb = load_workbook(CONFIG_PATH, data_only=True) # Cargar el libro de Excel con valores calculados
    return wb["CONFIG"] # Devolver la hoja "CONFIG"


# ---------------------------------------------------------
# Clase Vendor: representa un vendor y su estructura en el drive
# ---------------------------------------------------------
class Vendor:

    def get_base_path():
        ws = _load_config_sheet()
        value = ws["B3"].value

        if not value:
            raise ValueError("BASE_PATH no definido en CONFIG!B3")

        return value

    BASE_PATH = get_base_path()  # Ruta base donde se guardan los vendors destino G:\Unidades compartidas\Vendor_files

    def __init__(self, name):
        """
        Inicializa un nuevo vendor con su nombre y ruta base.
        """
        self.name = name                                     # Nombre del vendor (string)
        self.base_path = Vendor.BASE_PATH                    # Referencia a la ruta base (constante de clase)
        self.vendor_path = os.path.join(Vendor.BASE_PATH, name)  # Ruta completa del vendor (base + nombre)
        self.current_year = datetime.now().year              # Año actual (int)
        self.current_month = datetime.now().month            # Mes actual (int)
        self.notifier = NOTIFIER                             # Notificador Slack (puede ser None)


    # ---------------------------------------------------------
    # Crear carpeta principal del vendor
    # ---------------------------------------------------------
    def create_vendor(self):
        """
        Crea la carpeta principal del vendor si no existe y devuelve True si se creó.
        """
        if not os.path.exists(self.vendor_path):             # Si la carpeta del vendor no existe
            os.makedirs(self.vendor_path)                    # Crear la carpeta del vendor
            msg = f"Vendor '{self.name}' creado con éxito."  # Mensaje de éxito
            print(msg)                                       # Mostrar por consola
            return True                                      # Indicar que se creó

        else:                                                # Si la carpeta ya existía
            msg = f"⚠️ La carpeta del vendor '{self.name}' ya existe."  # Mensaje de aviso
            print(msg)                                       # Mostrar aviso por consola
            return False                                     # Indicar que no se creó


    # ---------------------------------------------------------
    # Crear carpeta del año actual
    # ---------------------------------------------------------
    def create_year_folder(self):
        """
        Crea la carpeta del año actual dentro del vendor si no existe y devuelve su ruta.
        """
        year_path = os.path.join(self.vendor_path, str(self.current_year))  # Ruta al folder del año dentro del vendor

        if not os.path.exists(year_path):                 # Si la carpeta del año no existe
            os.makedirs(year_path)                        # Crear carpeta del año
            print(f"📁 Carpeta creada para el año: {self.current_year}")  # Mensaje de creación
        else:
            print(f"✅ Carpeta del año {self.current_year} ya existe.")    # Mensaje si ya existía

        return year_path                                   # Devolver la ruta del año (creada o existente)


    # ---------------------------------------------------------
    # Crear estructura del mes actual
    # ---------------------------------------------------------
    def create_month_structure(self):
        """
        Crea la estructura del mes actual dentro del año correspondiente:
        - Carpeta "MM.MonthName"
        - Subcarpetas 'ordenes/OEA', 'ordenes/OE JR' y 'cierres'
        """
        year_path = self.create_year_folder()              # Asegurar que la carpeta del año exista y obtener su ruta

        month_folder = f"{self.current_month:02d}.{calendar.month_name[self.current_month]}"  # Nombre carpeta mes "MM.MonthName"
        month_path = os.path.join(year_path, month_folder)  # Ruta completa al folder del mes

        if not os.path.exists(month_path):                 # Si la carpeta del mes no existe
            os.makedirs(month_path)                        # Crear la carpeta del mes
            print(f"📁 Carpeta creada para el mes: {month_folder}")  # Mensaje de creación
        else:
            print(f"✅ Carpeta del mes {month_folder} ya existe.")    # Mensaje si ya existía

        # Rutas a subcarpetas dentro del mes
        orders_path = os.path.join(month_path, "ordenes")  # Ruta a la carpeta 'ordenes' dentro del mes
        closures_path = os.path.join(month_path, "cierres")# Ruta a la carpeta 'cierres' dentro del mes

        # Crear las subcarpetas de órdenes (exist_ok evita excepción si ya existen)
        os.makedirs(os.path.join(orders_path, "OEA"), exist_ok=True)   # 'ordenes/OEA'
        os.makedirs(os.path.join(orders_path, "OE JR"), exist_ok=True) # 'ordenes/OE JR'

        # Crear la carpeta de cierres (exist_ok evita excepción si ya existe)
        os.makedirs(closures_path, exist_ok=True)

        print("📂 Estructura mensual creada correctamente.")  # Confirmación final


    # ---------------------------------------------------------
    # Función principal para crear todo un vendor automaticamente
    # ---------------------------------------------------------
    def update_structure(self):
        """
        Verifica y crea la estructura completa del vendor:
        - Crea el vendor
        - Crea el año actual y la estructura del mes actual
        Envía mensajes por consola y utiliza notifier si está configurado.
        """
        vendor_created = self.create_vendor()                                # Crear carpeta principal del vendor
        self.create_month_structure()                                        # Crear año/mes/ordenes/cierres

        month_folder = f"{self.current_month:02d}.{calendar.month_name[self.current_month]}"  # Nombre del mes en formato MM.MonthName

        if vendor_created:                                                    # Si el vendor no existe
            msg = f"""🗂️ Vendor '{self.name}' creado con éxito:
        -    Estructura inicial creada para el año: {self.current_year}
        -    Mes: {month_folder}"""
            print(msg)                                                       # Mostrar mensaje resumen por consola
            if self.notifier:                                                 # Enviar por Slack el mismo mensaje
                self.notifier.send(msg)
        else:                                                                  # Si el vendor ya existía
            msg = f"""⚠️ Vendor '{self.name}' ya existía."""
            if self.notifier:                                                 # Enviar aviso por Slack
                self.notifier.send(msg)


    # ---------------------------------------------------------
    # Crear estructura de un mes y año personalizados en un solo vendor
    # ---------------------------------------------------------
    def create_custom_month_structure(self, vendor_name, year, month):
        """
        Crea la estructura de carpetas para un año y mes específicos.
        - Valida mes
        - Ajusta temporalmente current_year/current_month para reutilizar métodos existentes
        - Crea carpetas y subcarpetas debidas
        - Restaura valores originales
        """
        if not 1 <= month <= 12:                                   # Validar rango del mes
            print("❌ El mes debe estar entre 1 y 12.")             # Mensaje de error si el mes no está en 1..12
            return                                                 # Salir sin hacer cambios

        previous_year = self.current_year                          # Guardar valor actual de año
        previous_month = self.current_month                        # Guardar valor actual de mes
        self.current_year = year                                    # Sustituir temporalmente el año
        self.current_month = month                                  # Sustituir temporalmente el mes

        self.create_vendor()                                        # Crear carpeta principal del vendor si hace falta
        year_path = self.create_year_folder()                       # Crear/obtener carpeta del año modificado

        month_name = calendar.month_name[month]                     # Nombre del mes solicitado (ej. 'December')
        month_folder = f"{month:02d}.{month_name}"                  # Formato carpeta destino: "MM.MonthName"
        month_path = os.path.join(year_path, month_folder)          # Ruta completa al folder del mes

        if not os.path.exists(month_path):                          # Si la carpeta del mes no existe
            os.makedirs(month_path)                                 # Crear la carpeta del mes
            msg = f"""📁✅ Se ha creado carpeta para el vendor: {vendor_name}
            En el año: '{year}'
            En el mes: '{month_folder}' exitosamente."""
            print(msg)                                             # Mensaje de creación exitosa
            if self.notifier:                                      # Enviar notificación si está configurado
                self.notifier.send(msg)
        else:
            msg = f"⚠️ Carpeta del vendor {vendor_name} para el mes {month_folder} ya existe."
            print(msg)                                             # Mensaje si ya existía
            if self.notifier:
                self.notifier.send(msg)

        # Crear subcarpetas internas necesarias
        orders_path = os.path.join(month_path, "ordenes")           # Ruta a 'ordenes' dentro del mes
        closures_path = os.path.join(month_path, "cierres")         # Ruta a 'cierres' dentro del mes

        os.makedirs(os.path.join(orders_path, "OEA"), exist_ok=True)   # Crear 'ordenes/OEA'
        os.makedirs(os.path.join(orders_path, "OE JR"), exist_ok=True) # Crear 'ordenes/OE JR'
        os.makedirs(closures_path, exist_ok=True)                     # Crear 'cierres'

        print("📂 Estructura personalizada creada correctamente.")   # Confirmación final

        # Restaurar valores originales
        self.current_year = previous_year
        self.current_month = previous_month


# ---------------------------------------------------------
# Nueva clase para gestionar varios vendors (origen = Marketing Team)
# ---------------------------------------------------------
class ManagerVendors:

    def get_origin_path():
        ws = _load_config_sheet()
        value = ws["B2"].value

        if not value:
            raise ValueError("ORIGIN_PATH no definido en CONFIG!B2")

        return value

    BASE_PATH = get_origin_path()  # Ruta base origen (donde están los insertion orders) G:\Unidades compartidas\Marketing Team\Offline Marketing\03. Insertion orders\01. TV
    notifier = NOTIFIER   # Notificador de Slack opcional (puede ser None)


    @classmethod
    def update_all_vendors_month(cls, year, month):
        """
        Crea la estructura de un mes/año específicos para todos los vendors dentro de Vendor.BASE_PATH.
        """
        if not os.path.exists(Vendor.BASE_PATH):                       # Verificar que la ruta destino (vendors) exista
            print(f"❌ La ruta destino '{cls.BASE_PATH}' no existe. No se puede continuar.")  # Mensaje de error
            return                                                     # Salir si no existe

        for vendor_name in os.listdir(Vendor.BASE_PATH):               # Iterar sobre cada vendor en la ruta destino
            vendor_path = os.path.join(Vendor.BASE_PATH, vendor_name)  # Ruta al vendor
            if os.path.isdir(vendor_path):                             # Si la entrada es un directorio (vendor)
                vendor = Vendor(vendor_name)                           # Crear instancia temporal de Vendor
                try:
                    vendor.create_custom_month_structure(vendor_name, year, month)  # Crear estructura personalizada para ese vendor
                except Exception as e:
                    print(f"❌ Error al actualizar {vendor_name}: {e}")  # Informar error si falla
                else:
                    print(f"✅ Actualizado vendor: {vendor_name}")      # Informar éxito

        print("🎯 Estructura de mes/año creada para todos los vendors.")  # Mensaje final del proceso


    @staticmethod
    def normalize(text: str) -> str:
        """
        Normaliza texto para comparaciones: elimina espacios redundantes y pasa a minúsculas.
        """
        return text.strip().lower()


    @classmethod
    def copy_latest_order(cls, vendor_name: str, product: str, year: int, month: int):
        """
        Copia el último archivo de insertion orders desde la ruta ORIGEN (ManagerVendors.BASE_PATH)
        al vendor correspondiente dentro de Vendor.BASE_PATH, según product (OE / OE JR), año y mes.
        """
        # Imprimir parámetros recibidos para trazabilidad
        print("\n================ INICIO DEL PROCESO ================")
        print(f"> vendor_name = {vendor_name}")
        print(f"> product = {product}")
        print(f"> year = {year}")
        print(f"> month = {month}")
        print("====================================================\n")

        vendor_name_clean = cls.normalize(vendor_name)                # Normalizar vendor para comparación
        product_clean = cls.normalize(product)                        # Normalizar producto para comparación

        if product_clean not in ["oe", "oe jr"]:                      # Validar producto permitido
            print("❌ Error: producto inválido")                       # Informar error por producto inválido
            return                                                    # Salir sin más acciones

        origin_product_folder = "01. OE" if product_clean == "oe" else "02. OE JR"  # Carpeta origen según producto
        dest_product_folder = "OEA" if product_clean == "oe" else "OE JR"          # Carpeta destino según producto

        print(f"[INFO] Carpeta de producto origen esperada: {origin_product_folder}")  # Informar carpeta origen esperada
        print(f"[INFO] Carpeta de producto destino esperada: {dest_product_folder}\n") # Informar carpeta destino esperada

        # Indicar ruta origen desde donde se leerán los vendors/archivos
        print(f"[INFO] Leyendo carpetas en origen BASE_PATH:\n{cls.BASE_PATH}\n")

        file_found = False                                          # Flag para indicar si se encontró el vendor

        for vendor_folder in os.listdir(cls.BASE_PATH):                  # Iterar sobre cada carpeta en ruta origen
            vendor_path = os.path.join(cls.BASE_PATH, vendor_folder)     # Construir ruta completa al folder actual
            print(f"[CHECK] Revisando folder: {vendor_folder}")          # Mostrar carpeta que se está revisando

            if not os.path.isdir(vendor_path):                          # Si la entrada no es directorio, omitirla
                print("   - No es carpeta, se ignora.")
                continue

            # Buscar carpeta del año dentro del vendor origen (ej. "Año 2025")
            year_folder_name = f"Año {year}"                            # Nombre esperado de carpeta año en origen
            year_folder_path = os.path.join(vendor_path, year_folder_name)  # Ruta esperada al folder del año
            print(f"   > Buscando carpeta de año: {year_folder_name}")

            if not os.path.isdir(year_folder_path):                     # Si no existe la carpeta del año, continuar con siguiente vendor
                print("     ✖ No existe esta carpeta de año, continuar con siguiente vendor\n")
                continue

            print("     ✔ Carpeta de año encontrada.")                  # Confirmar que se encontró la carpeta del año

            # Buscar carpeta del mes en formato origen (prefijo "MM-")
            target_month_prefix = f"{month:02d}-"                       # Prefijo que identifica la carpeta del mes en origen
            print(f"   > Buscando carpeta de mes con prefijo: {target_month_prefix}")

            month_folder = None                                         # Inicializar variable para almacenar carpeta mes encontrada
            for folder in os.listdir(year_folder_path):                 # Iterar entradas dentro de la carpeta del año
                if cls.normalize(folder).startswith(cls.normalize(target_month_prefix)):  # Comparar prefijos normalizados
                    month_folder = folder                               # Guardar carpeta mes encontrada
                    break

            if not month_folder:                                        # Si no se encontró la carpeta del mes, continuar
                print("     ✖ No se encontró carpeta del mes.\n")
                continue

            print(f"     ✔ Carpeta de mes encontrada: {month_folder}")    # Mostrar carpeta mes encontrada

            month_folder_path = os.path.join(year_folder_path, month_folder)  # Ruta completa al folder del mes

            # Dentro del mes, buscar la carpeta del producto origen (01. OE o 02. OE JR)
            print(f"   > Buscando carpeta del producto: {origin_product_folder}")

            product_folder_path = os.path.join(month_folder_path, origin_product_folder)  # Ruta a carpeta de producto origen
            if not os.path.isdir(product_folder_path):                        # Si no existe esa carpeta, pasar al siguiente vendor
                print("     ✖ No existe carpeta de producto.\n")
                continue

            print("     ✔ Carpeta de producto encontrada.")                    # Confirmar carpeta producto

            # Entrar a '02. Insertion orders' dentro de la carpeta del producto
            print("   > Entrando a carpeta '02. Insertion orders'")

            insertion_orders_path = os.path.join(product_folder_path, "02. Insertion orders")  # Ruta a insertion orders
            if not os.path.isdir(insertion_orders_path):                         # Si no existe subcarpeta, continuar
                print("     ✖ No existe '02. Insertion orders'.\n")
                continue

            print("     ✔ Carpeta '02. Insertion orders' encontrada.")            # Confirmación existencia

            # Listar archivos en la carpeta de insertion orders (solo archivos, no directorios)
            print("   > Listando archivos en carpeta de inserción...")
            all_files = [
                f for f in os.listdir(insertion_orders_path)
                if os.path.isfile(os.path.join(insertion_orders_path, f))
                and not f.startswith(".")
                and not f.startswith("~$")
            ]

            print("     [DEBUG] Archivos detectados:")
            for f in all_files:
                print(f"         -> '{f}'")


            print(f"     ✔ {len(all_files)} archivos encontrados.")                # Mostrar cantidad de archivos encontrados

            if not all_files:                                                     # Si no hay archivos, continuar con siguiente vendor
                print("     ✖ No hay archivos, continuar.\n")
                continue

            # Regla de negocio: la 'versión' a copiar está basada en la cantidad de archivos
            total_files = len(all_files)                                          # Contar archivos disponibles
            version_to_copy = f"{total_files}."                                   # Prefijo de versión esperado en nombre de archivo (ej. "5.")

            print(f"   > Versión esperada a copiar: {version_to_copy}")            # Informar versión esperada

            # Buscar archivo que cumpla prefijo de versión y cuyo vendor coincida con el solicitado
            print("   > Buscando archivo que coincida con la versión y vendor...")
            file_to_copy = None                                                    # Inicializar archivo seleccionado
            for file in all_files:                                                 # Iterar candidatos
                if cls.normalize(file).startswith(cls.normalize(version_to_copy)): # Filtrar por prefijo de versión
                    print(f"     - Candidato encontrado: {file}")                  # Mostrar candidato detectado

                    parts = file.split()                                           # Separar nombre en palabras por espacios

                    try:
                        if product_clean == "oe":                                 # Para OE normal, vendor comienza en la posición esperada 4
                            vendor_in_file = " ".join(parts[4:])                   # Extraer vendor desde posición 4 en adelante
                        else:                                                     # Para OE JR, vendor comienza en posición 5
                            vendor_in_file = " ".join(parts[5:])                   # Extraer vendor desde posición 5 en adelante

                        vendor_in_file = os.path.splitext(vendor_in_file)[0]      # Quitar extensión para comparar solo el nombre
                    except IndexError:
                        print("       ✖ Error analizando nombre del archivo.")     # Manejo de nombres que no siguen la estructura esperada
                        continue                                                   # Saltar candidato inválido

                    print(f"       > Vendor extraído del archivo: {vendor_in_file}")  # Mostrar vendor extraído

                    if cls.normalize(vendor_in_file) == vendor_name_clean:          # Comparar vendor extraído con vendor solicitado
                        print("       ✔ Coincidencia encontrada con el vendor solicitado.")
                        file_to_copy = file                                           # Seleccionar este archivo para copiar
                        break                                                         # Romper búsqueda de archivos
                    else:
                        print("       ✖ Vendor no coincide.")                        # Indicar diferencia y continuar

            if not file_to_copy:                                                      # Si no se encontró un archivo coincidente
                print("     ✖ No se encontró un archivo que coincida con vendor y versión.\n")
                continue
            else:
                file_found = True                                               # Indicar que se encontró el vendor


            print(f"     ✔ Archivo final a copiar: {file_to_copy}\n")                   # Mostrar archivo seleccionado

            # Ruta completa del archivo origen seleccionado
            origin_file_path = os.path.join(insertion_orders_path, file_to_copy)      # Construir ruta del archivo origen
            print(f"[INFO] Ruta completa del archivo origen:\n{origin_file_path}\n")    # Mostrar ruta origen para trazabilidad

            # Buscar vendor destino equivalente dentro de Vendor.BASE_PATH
            print("\n[INFO] Buscando ruta destino en Vendor.BASE_PATH...\n")
            dest_base = Vendor.BASE_PATH                                              # Ruta base destino (vendors)
            print(f"[INFO] Ruta base destino: {dest_base}")

            dest_vendor_folder = None                                                  # Inicializar variable para carpeta destino encontrada
            for folder in os.listdir(dest_base):                                       # Iterar carpetas en ruta destino
                if cls.normalize(folder) == vendor_name_clean:                        # Comparar nombres normalizados
                    dest_vendor_folder = folder                                       # Guardar nombre real de carpeta destino
                    break

            if not dest_vendor_folder:                                                 # Si no se encontró carpeta destino, terminar con error
                print("❌ No existe carpeta destino del vendor.\n")
                return

            print(f"✔ Carpeta destino del vendor: {dest_vendor_folder}")                 # Informar carpeta destino encontrada

            dest_vendor_path = os.path.join(dest_base, dest_vendor_folder)              # Ruta completa al vendor destino

            # Buscar carpeta del año en el vendor destino
            dest_year_path = os.path.join(dest_vendor_path, str(year))                  # Ruta al folder del año destino dentro del vendor
            print(f"> Buscando carpeta año destino: {dest_year_path}")

            if not os.path.isdir(dest_year_path):                                       # Si no existe carpeta del año en destino, devolver error
                print("❌ No existe carpeta destino del año.\n")
                return f"❌ En el vendor '{vendor_name}' destino no existe la carpeta del año {year}.\n"

            print("✔ Carpeta año destino encontrada.")                                   # Confirmación año encontrado

            # Buscar carpeta del mes destino con formato "MM.MonthName"
            print("> Buscando carpeta de mes destino...")
            month_prefix_point = f"{month:02d}."                                        # Prefijo que identifica carpeta mes detino ("MM.")
            dest_month_folder = None                                                    # Inicializar variable para carpeta mes destino
            for folder in os.listdir(dest_year_path):                                   # Iterar carpetas dentro del año destino
                if cls.normalize(folder).startswith(cls.normalize(month_prefix_point)): # Buscar carpeta que empiece con "MM."
                    dest_month_folder = folder                                         # Guardar carpeta encontrada
                    break

            if not dest_month_folder:                                                    # Si no se encontró carpeta mes destino, devolver error
                print("❌ No existe carpeta destino del mes.\n")
                return f"❌ En el vendor '{vendor_name}' no existe el mes {month}"

            print(f"✔ Carpeta mes destino encontrada: {dest_month_folder}")               # Informar carpeta mes destino encontrada

            dest_month_path = os.path.join(dest_year_path, dest_month_folder)            # Ruta completa al folder del mes destino

            # Verificar carpeta 'ordenes' dentro del mes destino
            orders_path = os.path.join(dest_month_path, "ordenes")                       # Ruta a 'ordenes' dentro del mes destino
            print("> Buscando carpeta 'ordenes'...")

            if not os.path.isdir(orders_path):                                           # Si no existe carpeta 'ordenes', devolver error
                print("❌ No existe carpeta 'ordenes'.\n")
                return f"❌ En el vendor | {vendor_name} | año '{year}' | mes {month} | no existe la carpeta ordenes"

            print("✔ Carpeta 'ordenes' encontrada.")                                     # Confirmar existencia

            # Construir ruta final al producto dentro de 'ordenes'
            final_dest_path = os.path.join(orders_path, dest_product_folder)             # Ruta final donde se copiará el archivo
            print(f"> Buscando carpeta destino final: {final_dest_path}")

            if not os.path.isdir(final_dest_path):                                       # Si no existe carpeta de producto destino, devolver error
                print("❌ No existe carpeta final del producto.\n")
                return f"❌ En el vendor | {vendor_name} | mes {month} | no existe la carpeta '{product}' debes eliminar la carpeta de este mes y crearla de nuevo"

            print("✔ Carpeta final destino encontrada.\n")                               # Confirmación final

            # Copiar archivo desde origen a destino preservando metadatos
            print(">>> COPIANDO ARCHIVO...")
            shutil.copy2(origin_file_path, final_dest_path)                              # Copiar archivo (preserva metadatos)

            print(f"✅ Archivo copiado exitosamente a:\n{final_dest_path}")               # Confirmación de copia
            print("================ FIN DEL PROCESO ====================\n")              # Mensaje final de proceso
            return f"✅ El vendor '{vendor_name}' se ha actualizado correctamente de version\n -    Producto: {product}\n -    Año: {year}\n -    Mes {month}.\n"

        # Si se recorrió todo origen y no se encontró archivo coincidente para el vendor solicitado
        print("❌ No se encontró ningún archivo coincidente con el vendor solicitado.\n")
        if file_found is False:
            return f"❌ No se encontró ningun archivo en el vendor '{vendor_name}' del producto '{product}' en la ruta origen.\n"


    @classmethod
    def copy_latest_orders_batch(cls, vendors: list, product: str, year: int, month: int):
        """
        Ejecuta copy_latest_order para varios vendors (proceso por lotes).
        """
        print("\n=========== INICIO PROCESO POR LOTES ===========\n")  # Encabezado batch

        for vendor in vendors:                                      # Iterar lista de vendors proporcionada
            print(f"\n>>> Ejecutando para vendor: {vendor}")        # Mensaje por vendor
            print("---------------------------------------------")
            try:
                result = cls.copy_latest_order(vendor, product, year, month)  # Llamada al proceso para cada vendor
                print(result)                                               # Mostrar resultado de la operación
                if cls.notifier:                                             # Enviar resumen por Slack si está configurado
                    cls.notifier.send(result)
            except Exception as e:
                print(f"❌ Error inesperado con vendor {vendor}: {e}")        # Capturar e informar errores por vendor

        print("\n=========== FIN PROCESO POR LOTES ===========\n")     # Mensaje final batch


class ConsoleRedirect:
    """
    Redirige stdout/stderr hacia un wi
    dget Text de Tkinter (útil para integrar consola en GUI).
    """
    def __init__(self, text_widget):
        self.text_widget = text_widget   # Widget Text donde se escribirán las salidas

    def write(self, string):
        """
        Inserta texto en el widget Text respetando estado (readonly) y desplazando el scroll.
        """
        if not self.text_widget.winfo_exists():
            return 
        self.text_widget.config(state="normal")
        self.text_widget.insert("end", string)   # Insertar el texto recibido
        self.text_widget.see("end")              # Asegurar visibilidad del final del contenido
        self.text_widget.config(state="disabled")# Volver a dejar el widget en modo sólo lectura

    def flush(self):
        """Implementación vacía para compatibilidad con la interfaz de streams."""
        pass

    @staticmethod
    def attach(text_widget):
        """
        Atacha una instancia de ConsoleRedirect a sys.stdout y sys.stderr para redirección.
        Devuelve la instancia creada.
        """
        redirector = ConsoleRedirect(text_widget)  # Crear redirector con el widget dado
        sys.stdout = redirector                     # Redirigir stdout
        sys.stderr = redirector                     # Redirigir stderr
        return redirector

